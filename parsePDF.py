import pdfplumber
import re
import os
import openpyxl
from openpyxl import Workbook


##射频模块要求在前一页，表格在后一页最上面的位置

def searchPDF(path, wb, fileName):
    pdf = pdfplumber.open(path)
    pageNum = 0
    pages = pdf.pages

    resSearchNum = 0
    continuedPage = False  # 通过一个flag来判断table是否在下一页也是特性映射中的内容，flag不变，直到table检测不到“特性id”或者没有表格
    lastPageTableTitle = ""
    sheetIsExisted = False
    wb.create_sheet(fileName)
    titleFontSize = 0

    for page in pages:
        text = page.extract_text()

        resSearch = re.search("特性映射", text)
        resSearch2 = re.search("射频模块要求", text)

        pageNum += 1
        print("find %d time" % pageNum)

        # 看下一页的情况
        if (continuedPage == True):
            # 同一个表分在两页
            tables, continuedPage = writeContinuedPageIntoExcel(page, wb, fileName, titleFontSize)
            # tables[0] = tables[0][1:len(tables[0])]#去掉重复的标题
            lastPageTableTitle = ""

        # 找到了正文中的 特性映射
        if (resSearch is not None) and (resSearchNum == 1):
            startIdx = getCharIdx(page, "特性映射")  # 得到 “射”或者“特”的index
            titleFontSize = float(page.chars[startIdx]['adv'])
            tables, continuedPage = writeFirstPageIntoExcel(page, "特性映射", wb, fileName, titleFontSize)
            #if (tables != []):
                #print(tables)
            resSearchNum = 0

            # 第一次出现 特性映射 在标题中
        if (resSearch is not None):
            resSearchNum += 1

        # 找到了 射频模块要求
        if (resSearch2 is not None):

            # 打印 射频模块要求 所在的大标题
            startRow = wb[fileName].max_row + 1
            #print(getCurrPageLargeTitle(text))
            wb[fileName].cell(row=1 + startRow, column=1, value=getCurrPageLargeTitle(text))

            startIdx = getCharIdx(page, "射频模块要求")  # 得到 “射”或者“特”的index
            titleFontSize = float(page.chars[startIdx]['adv'])
            tables, continuedPage = writeFirstPageIntoExcel(page, "射频模块要求", wb, fileName, titleFontSize)
            #if (tables != []):
                #print(tables[0])
                #lastPageTableTitle = tables[0][0]


def findRegionOfTablesOrTexts(topLineHeight, page, titleFontSize, i):
    textRequirement = ""
    continuedPage = False
    j = i

    while (i < len(page.chars) and float(page.chars[i]['adv']) < titleFontSize):
        i += 1

    bottomCharHeight = page.chars[min(len(page.chars) - 1, i)]['top']  # 下一个标题或者页脚 所在行高

    # 文字描述：到下一个标题为止或者底部线以上为止
    while (float(page.chars[j]['adv']) < titleFontSize and float(page.chars[j]['y0']) > 67.8 and float(
            page.chars[j]['top']) > 85):  # and page.chars[j]['fontname'] != 'CANYFW+HuaweiSans-Bold'):
        textRequirement += page.chars[j]['text']
        j += 1

    # 找到页面底部 也没有大于标题的 字体,就说明还有下一页
    if i == len(page.chars):
        continuedPage = True
    # 找到比title大的字体
    else:
        charHeight = page.chars[i]['bottom'] #第一个大于titleFontSize的character

        i = i - 1
        for k in range(len(textRequirement) - 1, -1, -1):
            #print(page.chars[i]['bottom'])
            if charHeight - page.chars[i]['bottom'] <= 0.3:
                textRequirement = textRequirement[0 : -1]
            else:
                break
            i -= 1

    betweenTwoTitle = page.crop((0, topLineHeight, page.width, bottomCharHeight))
    tables = betweenTwoTitle.extract_tables()
    return tables, continuedPage, textRequirement


# 把找到关键字的那一页的相关信息写入excel表中
def writeFirstPageIntoExcel(page, target, wb, fileName, titleFontSize):
    startIdx = getCharIdx(page, target)
    i = startIdx + len(target)  # i：是对文本遍历（为了找到下一个标题）的开始的index
    topCharHeight = page.chars[startIdx]['bottom']  # 射频模块要求 所在行高
    tables, continuedPage, textRequirement = findRegionOfTablesOrTexts(topCharHeight, page, titleFontSize, i)
    writeTableOrTextIntoExcel(tables, wb, fileName, textRequirement)
    return tables, continuedPage


# 把找到关键字的后几页的相关信息写入excel表中，直到找到下一个的标题
def writeContinuedPageIntoExcel(page, wb, fileName, titleFontSize):
    topLineHeight = 85
    i = 0
    tables, continuedPage, textRequirement = findRegionOfTablesOrTexts(topLineHeight, page, titleFontSize, i)
    writeTableOrTextIntoExcel(tables, wb, fileName, textRequirement)
    return tables, continuedPage


def writeTableOrTextIntoExcel(tables, wb, fileName, textRequirement):
    if (tables != []):
        for table in tables:
            startRow = wb[fileName].max_row
            #print(table)
            writeTableIntoExcel(table, wb[fileName], startRow)
    else:
        startRow = wb[fileName].max_row
        wb[fileName].cell(row=1 + startRow, column=1, value=textRequirement)
        #print(textRequirement)


def findTargetStr(page, target, i):
    for targetChar in target:
        if (page.chars[i]['text'] != targetChar):
            return False
        i += 1
    return True


def getCurrPageLargeTitle(text):
    line = 0
    thirdLine = ""
    for ch in text:
        if (ch == '\n'):
            line += 1
            continue
        if (line == 1):
            thirdLine += ch
        elif (line > 1):
            break
    words = thirdLine.split()

    # 得到标题，通过找到章节的数字
    title = ""
    i = 0
    for i in range(0, len(words)):
        if (words[i].isdigit()):
            break

    for j in range(i + 1, len(words)):
        title += words[j]
    #print(title)
    return title


'''            
def searchSecondRequire(text):
    paragraphs = text.split("射频模块要求", 1)
    requirement = ""
    for ch in paragraphs[1]:
        if()
        requirement += ch

'''


# 得到 射频模块要求 所在index
def getCharIdx(page, target):
    i = 0
    for ch in page.chars:
        if (findTargetStr(page, target, i)):
            return i
        i += 1


def walkFile(file, wb):
    for root, dirs, files in os.walk(file):
        for f in files:
            searchPDF(os.path.join(root, f), wb, f)


def writeTableIntoExcel(table, workSheet, startRow):
    for i in range(len(table)):
        for j in range(0, len(table[i])):
            if (str(table[i][j]) == "None"):
                workSheet.cell(row=i + 1 + startRow, column=j + 1, value=str(table[i - 1][j]))
                table[i][j] = table[i - 1][j]
            else:
                workSheet.cell(row=i + 1 + startRow, column=j + 1, value=str(table[i][j]))


# as long as table has TeXingID, return ture
def isTheTable(table, targetString):
    for row in range(len(table)):
        for col in range(len(table[0])):
            if (table[row][col] == targetString):
                return True
    return False


def main():
    #filefullpath = 'C:/Users/q50012319/AppData/Local/Programs/Python/Python37/Scripts/result/data.xlsx'
    currPath = os.getcwd()
    #if os.path.exists(filefullpath):
        #os.remove(filefullpath)
    wb = Workbook()
    walkFile(currPath+"/allPDFs", wb)
    wb.remove(wb["Sheet"])
    wb.save(currPath+'/resultData.xlsx')


if __name__ == '__main__':
    main()

